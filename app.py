import os
import sys
import tkinter as tk
from tkinter import messagebox
import openpyxl
import pandas as pd
from tkinter import ttk
import numpy as np
import joblib


def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
# Load the model
xgb_model = joblib.load(resource_path('data//saas_price_predictor_xgb.pkl'))

# Function to handle login
def login():
    security_value = None
    username = entry_username.get().strip()
    password = entry_password.get().strip()
    login_type = selected_value_login.get().strip()
    # Load the Excel data
    if login_type == "Customer":
        df = pd.read_excel(resource_path('data//customer_data.xlsx'))
    
    # Ensure data from Excel is also stripped of whitespace and converted to strings
        df.iloc[:, 0] = df.iloc[:, 0].astype(str).str.strip()  # Assuming username is in the first column
        df.iloc[:, 1] = df.iloc[:, 1].astype(str).str.strip()  # Assuming password is in the second column
        
        # Find matching rows
        matching_rows = df.loc[(df.iloc[:, 0] == username) & (df.iloc[:, 1] == password)]

        # Check if the user exists
        if not matching_rows.empty:
            security_value = True
        else:
            security_value = False
        
        # Display the result
        if security_value:
            # messagebox.showinfo("Login", "Login Successful!")
            show_dashboard_customer(username)
        else:
            messagebox.showerror("Login", "Incorrect username or password")

    else:
        df = pd.read_excel(resource_path('data//company_data.xlsx'))
    
    # Ensure data from Excel is also stripped of whitespace and converted to strings
        df.iloc[:, 0] = df.iloc[:, 0].astype(str).str.strip()  # Assuming username is in the first column
        df.iloc[:, 1] = df.iloc[:, 1].astype(str).str.strip()  # Assuming password is in the second column
        
        # Find matching rows
        matching_rows = df.loc[(df.iloc[:, 0] == username) & (df.iloc[:, 1] == password)]

        # Check if the user exists
        if not matching_rows.empty:
            security_value = True
        else:
            security_value = False
        
        # Display the result
        if security_value:
            # messagebox.showinfo("Login", "Login Successful!")
            show_dashboard_company(username)
        else:
            messagebox.showerror("Login", "Incorrect username or password")

# Function to show the dashboard in the same window
def show_dashboard_customer(username, start_number=0):
    # Hide all login and registration elements
    label_username.place_forget()
    entry_username.place_forget()
    label_password.place_forget()
    entry_password.place_forget()
    button_login.place_forget()
    text_name.place_forget()
    text_label.place_forget()
    dropdown_login.place_forget()
    label_user_type_login.place_forget()
    label_new_username.place_forget()
    entry_new_username.place_forget()
    label_new_password.place_forget()
    entry_new_password.place_forget()
    label_confirm_password.place_forget()
    entry_confirm_password.place_forget()
    button_register.place_forget()
    text_name_login.place_forget()
    text_label_login.place_forget()
    
    # Display the dashboard
    # dashboard_label = tk.Label(window, text=f"Welcome to Dashboard, {username}!", font=("Helvetica", 16))
    # dashboard_label.place(x=200, y=200)

    blue_band = tk.Frame(window, bg="blue", width=800, height=100)
    blue_band.place(x=0, y=0)

    # Add welcome text to the blue band
    welcome_label = tk.Label(blue_band, text="Price Predictor", font=("Helvetica", 20), bg="blue", fg="white")
    welcome_label.place(x=10, y=30)

    gray_band = tk.Frame(window, bg="gray", width=800, height=50)
    gray_band.place(x=0, y=100)

    df_customer_name = pd.read_excel(resource_path('data//customer_data.xlsx'))
    row = df_customer_name[df_customer_name['Username'] == username]
    user_real_name = row.iat[0, row.columns.get_loc('Name')]

    company_label = tk.Label(gray_band, text=f"Welcome to Dashboard, {user_real_name}", font=("Helvetica", 14), bg="gray", fg="black")
    company_label.place(x=10, y=10)

    product_label_band = tk.Frame(window, bg="white", width=700, height=35)
    product_label_band.place(x=0, y=170)
    product_info_label = tk.Label(product_label_band, text="Available Products", font=("Helvetica", 10), bg='white', fg="black")
    product_info_label.place(relx=0.5,rely=0.5, anchor='center')
    
    df_products = pd.read_excel(resource_path('data//product_data.xlsx'))
    product_list = df_products["Product Name"]
    company_list = df_products["Company Name"]
    price_list = df_products["Price"]
    discount_list = df_products["Discount"]
    start_number_history = start_number
    if len(product_list) != 0:
        last_number = start_number + 3
        while start_number < len(product_list) and start_number < last_number:
            product_label_band = tk.Frame(window, bg="white", width=150, height=180)
            product_label_band.place(x=50 + (start_number*225), y=260)

            product_name_band = tk.Frame(product_label_band, bg="lightblue", width=150, height=20)
            product_name_band.place(relx=0.5, y=40, anchor='center')
            product_label = tk.Label(product_name_band, text=str(product_list[start_number]), font=("Helvetica", 12), bg='white', fg="black")
            product_label.place(relx=0.5, rely=0.5, anchor='center')
            
            discount_label = tk.Label(product_label_band, text=f"({str(discount_list[start_number])} % off)", font=("Helvetica", 12), bg='white', fg="black")
            discount_label.place(relx=0.5, y=80, anchor='center')

            cost_label = tk.Label(product_label_band, text="Product Cost", font=("Helvetica", 9), bg='white', fg="black")
            cost_label.place(relx=0.5, y=105, anchor='center')
            discount_value = float(str(discount_list[start_number]))/100
            price_customer = round(float(str(price_list[start_number]))*(1-discount_value), 2)
            price_label = tk.Label(product_label_band, text=str(price_customer), font=("Helvetica", 14), bg='white', fg="green")
            price_label.place(relx=0.5, y = 130, anchor='center')

            start_number += 1
    else:
        last_number = 0
        product_list = 0

    if product_list != 0:
        if start_number_history < len(product_list):
            rate_label_1 = tk.Label(window, text="Rate Your Experience", bg='white', fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
            rate_label_1.bind("<Button-1>", lambda event: show_rate_window(start_number_history))
            rate_label_1.place(x=125, y=420, anchor='center')

        if start_number_history + 1 < len(product_list):
            rate_label_2 = tk.Label(window, text="Rate Your Experience", bg='white', fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
            rate_label_2.bind("<Button-1>", lambda event: show_rate_window(start_number_history+1))
            rate_label_2.place(x=350, y=420, anchor='center')

        if start_number_history + 2 < len(product_list):
            rate_label_3 = tk.Label(window, text="Rate Your Experience", bg='white', fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
            rate_label_3.bind("<Button-1>", lambda event: show_rate_window(start_number_history+2))
            rate_label_3.place(x=575, y=420, anchor='center')

    next_label = tk.Label(window, text="Next Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    next_label.bind("<Button-1>", lambda event:reload_dashboard_customer(username, start_number=last_number))

    previous_label = tk.Label(window, text="Previous Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    previous_label.bind("<Button-1>", lambda event:reload_dashboard_customer(username, start_number=last_number-6))
    if last_number > 3:
        previous_label.place(relx=0.1, y=480, anchor='center')

    if product_list != 0:
        if last_number < len(product_list):
            next_label.place(relx=0.9, y=480, anchor='center')

def show_rate_window(start_number):
    rate_window = tk.Toplevel(window)
    rate_window.title("Scrollable New Window")
    rate_window.geometry("350x250")

    rate_band = tk.Frame(rate_window, bg="lightblue", width=350, height=70)
    rate_band.place(relx=0.5, y=40, anchor='center')

    rate_label = tk.Label(rate_band, text="Rate Product", font=("Helvetica", 14), bg='lightblue', fg="black")
    rate_label.place(relx=0.5, rely=0.5, anchor='center')

    product_name_frame = tk.Frame(rate_window, bg="white", width=350, height=25)
    product_name_frame.place(relx=0.5, y=120, anchor='center')

    df_products_dummy = pd.read_excel(resource_path('data//product_data.xlsx'))
    product_list_dummy = df_products_dummy["Product Name"]
    product_name = product_list_dummy[start_number]
    product_name_label = tk.Label(product_name_frame, text=str(product_name), font=("Helvetica", 14), fg="black")
    product_name_label.place(relx=0.5, rely=0.5, anchor='center')

    instruction_label = tk.Label(rate_window, text="Rate between 1 to 5", font=("Helvetica", 10), fg="black")
    instruction_label.place(relx=0.5, y=150, anchor='center')

    # TODO: got another product (price high, found a better product)
    entry_rating = tk.Entry(rate_window)
    entry_rating.place(relx=0.5, y=180, anchor='center')
    button_rate = tk.Button(rate_window, text="Submit", command = lambda: get_rating(product_name, entry_rating, start_number, rate_window))
    button_rate.place(relx=0.5, y=220, anchor='center')

def get_rating(product_name, entry_rating, start_number, rate_window):
    rate_value = entry_rating.get().strip()
    rating_value = rate_value
    df_product_1 = pd.read_excel(resource_path('data//product_data.xlsx'))
    index = df_product_1.index[df_product_1['Product Name'] == product_name].tolist()
    
    row = df_product_1[df_product_1['Product Name'] == product_name]
    value = row.iat[0, row.columns.get_loc('Rating')]
    
    discount_value = row.iat[0, row.columns.get_loc('Discount')]

    if pd.isna(value) or value == "":
        rate_value = [rate_value]
        df_product_1.iloc[index, 4] = str(rate_value)
    else:
        value = eval(value)
        value += [rate_value]
        if len(value) % 5 == 0:
            int_value = [int(item) for item in value]
            rating_avg = sum(int_value)/len(int_value)
        if rating_avg < 2.5:
            if int(discount_value) < 35:
                discount_value = int(discount_value) + 5
                df_product_1.iloc[index, 3] = str(discount_value)

        df_product_1.iloc[index, 4] = str(value)

    df_product_1.to_excel(resource_path('data//product_data.xlsx'), index=False)
    
    if int(rating_value) < 3:
        reload_rating_window(start_number, rate_window)
    else:
        rate_window.destroy()

def reload_rating_window(start_number, rate_window):
    for widget in rate_window.winfo_children():
        widget.destroy()

    rate_band = tk.Frame(rate_window, bg="lightblue", width=350, height=70)
    rate_band.place(relx=0.5, y=40, anchor='center')

    rate_label = tk.Label(rate_band, text="Rate Product", font=("Helvetica", 14), bg='lightblue', fg="black")
    rate_label.place(relx=0.5, rely=0.5, anchor='center')

    product_name_frame = tk.Frame(rate_window, bg="white", width=350, height=25)
    product_name_frame.place(relx=0.5, y=120, anchor='center')

    df_products_dummy = pd.read_excel(resource_path('data//product_data.xlsx'))
    product_list_dummy = df_products_dummy["Product Name"]
    product_name = product_list_dummy[start_number]
    product_name_label = tk.Label(product_name_frame, text=str(product_name), font=("Helvetica", 14), fg="black")
    product_name_label.place(relx=0.5, rely=0.5, anchor='center')

    instruction_label = tk.Label(rate_window, text="Please select an reason", font=("Helvetica", 10), fg="black")
    instruction_label.place(relx=0.5, y=150, anchor='center')

    # TODO: got another product (price high, found a better product)
    selected_value_rating = tk.StringVar()
    selected_value_rating.set("Price is high")
    options_rating = ["Price is high", "Don't prefer to say"]
    dropdown_rating = tk.OptionMenu(rate_window, selected_value_rating, *options_rating)
    dropdown_rating.place(relx=0.5, y=180, anchor='center')
    button_rate = tk.Button(rate_window, text="Submit", command = lambda: get_reason(selected_value_rating, product_name, rate_window))
    button_rate.place(relx=0.5, y=220, anchor='center')

def get_reason(selected_value_rating, product_name, rate_window):
    rate_value = selected_value_rating.get().strip()
    df_product_1 = pd.read_excel(resource_path('data//product_data.xlsx'))
    index = df_product_1.index[df_product_1['Product Name'] == product_name].tolist()
    
    row = df_product_1[df_product_1['Product Name'] == product_name]
    value = row.iat[0, row.columns.get_loc('Reason')]
    
    if pd.isna(value) or value == "":
        rate_value = [rate_value]
        df_product_1.iloc[index, 5] = str(rate_value)
    else:
        value = eval(value)
        value += [rate_value]
        df_product_1.iloc[index, 5] = str(value)

    df_product_1.to_excel(resource_path('data//product_data.xlsx'), index=False)
    rate_window.destroy()

def reload_dashboard_customer(username, start_number):
    for widget in window.winfo_children():
        widget.destroy()

    blue_band = tk.Frame(window, bg="blue", width=800, height=100)
    blue_band.place(x=0, y=0)

    # Add welcome text to the blue band
    welcome_label = tk.Label(blue_band, text="Price Predictor", font=("Helvetica", 20), bg="blue", fg="white")
    welcome_label.place(x=10, y=30)

    gray_band = tk.Frame(window, bg="gray", width=800, height=50)
    gray_band.place(x=0, y=100)

    df_customer_name = pd.read_excel(resource_path('data//customer_data.xlsx'))
    row = df_customer_name[df_customer_name['Username'] == username]
    user_real_name = row.iat[0, row.columns.get_loc('Name')]

    company_label = tk.Label(gray_band, text=f"Welcome to Dashboard, {user_real_name}", font=("Helvetica", 14), bg="gray", fg="black")
    company_label.place(x=10, y=10)

    product_label_band = tk.Frame(window, bg="white", width=700, height=35)
    product_label_band.place(x=0, y=170)
    product_info_label = tk.Label(product_label_band, text="Available Products", font=("Helvetica", 10), bg='white', fg="black")
    product_info_label.place(relx=0.5,rely=0.5, anchor='center')
    
    df_products = pd.read_excel(resource_path('data//product_data.xlsx'))
    product_list = df_products["Product Name"]
    company_list = df_products["Company Name"]
    price_list = df_products["Price"]
    discount_list = df_products["Discount"]
    start_number_history = start_number
    if len(product_list) != 0:
        last_number = start_number + 3
        while start_number < len(product_list) and start_number < last_number:
            product_label_band = tk.Frame(window, bg="white", width=150, height=180)
            product_label_band.place(x=50 + ((start_number%3)*225), y=260)

            product_name_band = tk.Frame(product_label_band, bg="lightblue", width=150, height=20)
            product_name_band.place(relx=0.5, y=40, anchor='center')
            product_label = tk.Label(product_name_band, text=str(product_list[start_number]), font=("Helvetica", 12), bg='white', fg="black")
            product_label.place(relx=0.5, rely=0.5, anchor='center')
            
            discount_label = tk.Label(product_label_band, text=f"({str(discount_list[start_number])} % off)", font=("Helvetica", 12), bg='white', fg="black")
            discount_label.place(relx=0.5, y=80, anchor='center')

            cost_label = tk.Label(product_label_band, text="Product Cost", font=("Helvetica", 9), bg='white', fg="black")
            cost_label.place(relx=0.5, y=105, anchor='center')
            discount_value = float(str(discount_list[start_number]))/100
            price_customer = round(float(str(price_list[start_number]))*(1-discount_value), 2)
            price_label = tk.Label(product_label_band, text=str(price_customer), font=("Helvetica", 14), bg='white', fg="green")
            price_label.place(relx=0.5, y = 130, anchor='center')

            start_number += 1

    if start_number_history < len(product_list):
        rate_label_1 = tk.Label(window, text="Rate Your Experience", bg='white', fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
        rate_label_1.bind("<Button-1>", lambda event: show_rate_window(start_number_history))
        rate_label_1.place(x=125, y=420, anchor='center')

    if start_number_history + 1 < len(product_list):
        rate_label_2 = tk.Label(window, text="Rate Your Experience", bg='white', fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
        rate_label_2.bind("<Button-1>", lambda event: show_rate_window(start_number_history+1))
        rate_label_2.place(x=350, y=420, anchor='center')

    if start_number_history + 2 < len(product_list):
        rate_label_3 = tk.Label(window, text="Rate Your Experience", bg='white', fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
        rate_label_3.bind("<Button-1>", lambda event: show_rate_window(start_number_history+2))
        rate_label_3.place(x=575, y=420, anchor='center')
    
    next_label = tk.Label(window, text="Next Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    next_label.bind("<Button-1>", lambda event:reload_dashboard_customer(username, start_number=last_number))

    previous_label = tk.Label(window, text="Previous Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    previous_label.bind("<Button-1>", lambda event:reload_dashboard_customer(username, start_number=last_number-6))
    if last_number > 3:
        previous_label.place(relx=0.1, y=480, anchor='center')

    if last_number < len(product_list):
        next_label.place(relx=0.9, y=480, anchor='center')

def show_dashboard_company(username, start_number=0):
    # Hide all login and registration elements
    label_username.place_forget()
    entry_username.place_forget()
    label_password.place_forget()
    entry_password.place_forget()
    button_login.place_forget()
    text_name.place_forget()
    text_label.place_forget()
    dropdown_login.place_forget()
    label_user_type_login.place_forget()
    label_new_username.place_forget()
    entry_new_username.place_forget()
    label_new_password.place_forget()
    entry_new_password.place_forget()
    label_confirm_password.place_forget()
    entry_confirm_password.place_forget()
    button_register.place_forget()
    text_name_login.place_forget()
    text_label_login.place_forget()
    
    # Display the dashboard
    # dashboard_label = tk.Label(window, text=f"Welcome to Dashboard, {username}!", font=("Helvetica", 16))
    # dashboard_label.place(x=200, y=200)
    df_company_product = pd.read_excel(resource_path('data//company_data.xlsx'))
    row = df_company_product[df_company_product['Username'] == username]
    company_name = row.iat[0, row.columns.get_loc('Company')]
    blue_band = tk.Frame(window, bg="blue", width=800, height=100)
    blue_band.place(x=0, y=0)
    df_product_price = pd.read_excel(resource_path('data//product_data.xlsx'))

    # Add welcome text to the blue band
    welcome_label = tk.Label(blue_band, text="Price Predictor", font=("Helvetica", 20), bg="blue", fg="white")
    welcome_label.place(x=10, y=30)

    gray_band = tk.Frame(window, bg="gray", width=800, height=50)
    gray_band.place(x=0, y=100)

    company_label = tk.Label(gray_band, text=f"Welcome to Dashboard, {company_name}", font=("Helvetica", 14), bg="gray", fg="black")
    company_label.place(x=10, y=10)
    
    clickable_frame = tk.Frame(window, bg="lightblue", width=500, height=50)
    clickable_frame.place(x=100, y=170)

    start_number_history = start_number
    # Bind the frame to a click event
    clickable_frame.bind("<Button-1>", lambda e: open_new_window(username, start_number_history, company_name))

    frame_label = tk.Label(clickable_frame, text="Add New Product", bg="lightblue")
    frame_label.place(relx=0.5,rely=0.5, anchor='center')

    product_label_band = tk.Frame(window, bg="white", width=700, height=35)
    product_label_band.place(x=0, y=240)
    product_info_label = tk.Label(product_label_band, text="Product List", font=("Helvetica", 10), bg='white', fg="black")
    product_info_label.place(relx=0.5,rely=0.5, anchor='center')

    df_company_product = pd.read_excel(resource_path('data//company_data.xlsx'))
    row = df_company_product[df_company_product['Username'] == username]
    value = row.iat[0, row.columns.get_loc('Product')]
    if pd.isna(value) or value == "":
        no_product_label = tk.Label(window, text="No Registerd Products", font=("Helvetica", 10), fg="black")
        no_product_label.place(relx=0.5, y=370, anchor='center')
        last_number = start_number + 3
        product_data = 0
    else:
        last_number = start_number + 3
        product_data = eval(value)
        keys = product_data.keys()
        keys_list = list(keys)
        keys_list = [str(items) for items in keys_list]
        while start_number < len(product_data) and start_number < last_number:
            product_label_band = tk.Frame(window, bg="white", width=150, height=180)
            product_label_band.place(x=50 + (start_number*225), y=290)

            product_name_band = tk.Frame(product_label_band, bg="lightblue", width=150, height=20)
            product_name_band.place(relx=0.5, y=40, anchor='center')

            row_price = df_product_price[df_product_price['Product Name'] == keys_list[start_number]]
            value_price = row_price.iat[0, row_price.columns.get_loc('Price')]
            original_label = tk.Label(product_label_band, text="Original Cost", font=("Helvetica", 9), bg='white', fg="black")
            original_label.place(relx=0.5, y=135, anchor='center')

            asking_label = tk.Label(product_label_band, text="Asking Cost", font=("Helvetica", 9), bg='white', fg="black")
            asking_label.place(relx=0.5, y=80, anchor='center')
            asking_price = round(float(str(value_price)), 2)
            value_price = round(float(str(value_price))*0.5, 2)
            product_label = tk.Label(product_name_band, text=keys_list[start_number], font=("Helvetica", 12), bg='white', fg="black")
            product_label.place(relx=0.5, rely=0.5, anchor='center')

            price_label = tk.Label(product_label_band, text=str(value_price), font=("Helvetica", 14), bg='white', fg="green")
            price_label.place(relx=0.5, y = 160, anchor='center')

            asking_price_label = tk.Label(product_label_band, text=str(asking_price), font=("Helvetica", 14), bg='white', fg="lightblue")
            asking_price_label.place(relx=0.5, y = 105, anchor='center')
            start_number += 1

    next_label = tk.Label(window, text="Next Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    next_label.bind("<Button-1>", lambda event:reload_company_dashboard(username, start_number=last_number))

    previous_label = tk.Label(window, text="Previous Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    previous_label.bind("<Button-1>", lambda event:reload_dashboard_customer(username, start_number=last_number-6))
    if last_number > 3:
        previous_label.place(relx=0.1, y=485, anchor='center')
    if product_data != 0:   
        if last_number < len(product_data):
            next_label.place(relx=0.9, y=485, anchor='center')

def reload_company_dashboard(username, start_number):
    for widget in window.winfo_children():
        widget.destroy()

    df_company_product = pd.read_excel(resource_path('data//company_data.xlsx'))
    row = df_company_product[df_company_product['Username'] == username]
    company_name = row.iat[0, row.columns.get_loc('Company')]

    df_product_price = pd.read_excel(resource_path('data//product_data.xlsx'))
    start_number_history = start_number
    blue_band = tk.Frame(window, bg="blue", width=800, height=100)
    blue_band.place(x=0, y=0)

    # Add welcome text to the blue band
    welcome_label = tk.Label(blue_band, text="Price Predictor", font=("Helvetica", 20), bg="blue", fg="white")
    welcome_label.place(x=10, y=30)

    gray_band = tk.Frame(window, bg="gray", width=800, height=50)
    gray_band.place(x=0, y=100)

    company_label = tk.Label(gray_band, text=f"Welcome to Dashboard, {company_name}", font=("Helvetica", 14), bg="gray", fg="black")
    company_label.place(x=10, y=10)
    
    clickable_frame = tk.Frame(window, bg="lightblue", width=500, height=50)
    clickable_frame.place(x=100, y=170)

    # Bind the frame to a click event
    clickable_frame.bind("<Button-1>", lambda e: open_new_window(username, start_number_history, company_name))

    frame_label = tk.Label(clickable_frame, text="Add New Product", bg="lightblue")
    frame_label.place(relx=0.5,rely=0.5, anchor='center')

    product_label_band = tk.Frame(window, bg="white", width=700, height=35)
    product_label_band.place(x=0, y=240)
    product_info_label = tk.Label(product_label_band, text="Product List", font=("Helvetica", 10), bg='white', fg="black")
    product_info_label.place(relx=0.5,rely=0.5, anchor='center')

    df_company_product = pd.read_excel(resource_path('data//company_data.xlsx'))
    row = df_company_product[df_company_product['Username'] == username]
    value = row.iat[0, row.columns.get_loc('Product')]
    if pd.isna(value) or value == "":
        no_product_label = tk.Label(window, text="No Registerd Products", font=("Helvetica", 10), fg="black")
        no_product_label.place(relx=0.5, y=370, anchor='center')
    else:
        last_number = start_number + 3
        product_data = eval(value)
        keys = product_data.keys()
        keys_list = list(keys)
        keys_list = [str(items) for items in keys_list]
        while start_number < len(product_data) and start_number < last_number:
            product_label_band = tk.Frame(window, bg="white", width=150, height=180)
            product_label_band.place(x=50 + ((start_number%3)*225), y=290)
            
            product_name_band = tk.Frame(product_label_band, bg="lightblue", width=150, height=20)
            product_name_band.place(relx=0.5, y=40, anchor='center')
            
            row_price = df_product_price[df_product_price['Product Name'] == keys_list[start_number]]
            value_price = row_price.iat[0, row_price.columns.get_loc('Price')]
            original_label = tk.Label(product_label_band, text="Original Cost", font=("Helvetica", 9), bg='white', fg="black")
            original_label.place(relx=0.5, y=135, anchor='center')

            asking_label = tk.Label(product_label_band, text="Asking Cost", font=("Helvetica", 9), bg='white', fg="black")
            asking_label.place(relx=0.5, y=80, anchor='center')
            asking_price = round(float(str(value_price)), 2)
            value_price = round(float(str(value_price))*0.5, 2)
            product_label = tk.Label(product_name_band, text=keys_list[start_number], font=("Helvetica", 12), bg='white', fg="black")
            product_label.place(relx=0.5, rely=0.5, anchor='center')

            product_label = tk.Label(product_name_band, text=keys_list[start_number], font=("Helvetica", 12), bg='white', fg="black")
            product_label.place(relx=0.5, rely=0.5, anchor='center')

            price_label = tk.Label(product_label_band, text=str(value_price), font=("Helvetica", 14), bg='white', fg="green")
            price_label.place(relx=0.5, y = 160, anchor='center')

            asking_price_label = tk.Label(product_label_band, text=str(asking_price), font=("Helvetica", 14), bg='white', fg="lightblue")
            asking_price_label.place(relx=0.5, y = 105, anchor='center')
            start_number += 1

    next_label = tk.Label(window, text="Next Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    next_label.bind("<Button-1>", lambda event: reload_company_dashboard(username, start_number=last_number))

    previous_label = tk.Label(window, text="Previous Page", fg="blue", cursor="hand2", font=("Helvetica", 6, "underline"))
    previous_label.bind("<Button-1>", lambda event:reload_company_dashboard(username, start_number=last_number-6))
    if last_number > 3:
        previous_label.place(relx=0.1, y=485, anchor='center')
        
    if last_number < len(product_data):
        next_label.place(relx=0.9, y=485, anchor='center')

def open_new_window(username, start_number, company_name):
    # Create a new window
    new_window = tk.Toplevel(window)
    new_window.title("Scrollable New Window")
    new_window.geometry("350x500")

    # Create a canvas
    canvas = tk.Canvas(new_window)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Add a scrollbar to the canvas
    scrollbar = ttk.Scrollbar(new_window, orient="vertical", command=canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Create a frame within the canvas
    scrollable_frame = tk.Frame(canvas)

    # Configure the canvas to scroll with the scrollbar
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Enable scrolling with the mouse wheel
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    new_window.bind_all("<MouseWheel>", _on_mousewheel)  # For Windows and Linux
    new_window.bind_all("<Button-4>", _on_mousewheel)  # For Linux (some versions)
    new_window.bind_all("<Button-5>", _on_mousewheel)  # For Linux (some versions)
    new_window.bind_all("<Shift-MouseWheel>", lambda e: canvas.xview_scroll(int(-1 * (e.delta / 120)), "units"))

    # Populate the scrollable frame with data
    labels = ['Product Name', 'Revenue Multiple', 'Total Annual Revenue',
       'Team', 'Revenue',
       'Annual Profit', 'Annual Growth', 'Weekly Views', 'Business Age','Customers',
       'Listing Type','Location', 'What best describes the product', 'Tech Stack']
    
    raw_labels = ['productName', 'revenueMultiple','totalRevenueAnnual',
                  'team','revenue','annualProfit','growthAnnual',
                  'weeklyViews', 'businessAge', 'customers','listingType','location','keywords',
                  'techStack']
    
    entry_names = [i+"_entry" for i in labels[:9]]
    dropdown_vars = []
    for i in range(len(labels[9:])):
        dropdown_vars.append(tk.StringVar())

    dropdown_vars_name = [i+'_value' for i in labels[9:]]

    option_list = [['10-100', '100-1000', '1000-10000', '10000-100000', 'more than 100000'], ["platinum", "premium"], ["United States", "United Kingdom", "Canada", "France", "Switzerland", "Estonia", "Ukraine", "Spain", "Australia", "Germany", "Austria", "Latvia", "Belgium", "Turkey", "Mexico", "Singapore", "India", "Argentina"],
                   ["Automation", "Web Tools", "Analytics", "Technology", "Management", "HealthCare", "Mobile App", "Sales", "Finance", "SEO", "FundRaiser", "Workforce", "Servers", "Entertainment", "E-Commerce", "Loyalty", "Games", "Enterprise", "Marketing", "Security", "EdTech", "Telecommunication", "HRTech"],
                   ["FullStack and Cloud", "BackEnd and Cloud", "Cloud", "BackEnd", "FullStack", "FrontEnd", "FullStack and DBMS", "FrontEnd and Cloud", "Backend and DBMS", "Cloud and DBMS", "CyberSecurity", "DBMS","Not Listed"]]
    
    option_count = 0
    for items in dropdown_vars:
        items.set(option_list[option_count][0])
        option_count += 1

    dropdown_menu_vars = [i + "_menu" for i in dropdown_vars_name]
    scroll_frame_label = tk.Frame(scrollable_frame, bg="lightblue", width=330, height=50)
    scroll_frame_label.pack(pady=3, padx=10, anchor='w')
    
    upper_label = tk.Label(scroll_frame_label, text='Register Product', bg="lightblue", font=("Helvetica", 14))
    # upper_label.pack(padx=20, pady=20)
    upper_label.place(relx=0.5, rely=0.5, anchor='center')

    for i in range(len(labels[:9])):  # Example: adding 50 labels as data
        label = tk.Label(scrollable_frame, text=labels[i] + ':', font=("Helvetica", 10))
        label.pack(pady=3, padx=10, anchor='w')
        entry_names[i] = tk.Entry(scrollable_frame, width=50)
        entry_names[i].pack(pady=0, padx=10, anchor='w')

    for i in range(len(labels[9:])):
        label = tk.Label(scrollable_frame, text=labels[i+9] + ':', font=("Helvetica", 10))
        label.pack(pady=3, padx=10, anchor='w')
        dropdown_menu_vars[i] = tk.OptionMenu(scrollable_frame, dropdown_vars[i], *option_list[i])
        dropdown_menu_vars[i].config(width=45, anchor='w')
        dropdown_menu_vars[i].pack(pady=0, padx=10, anchor='w')

    button_register_product = tk.Button(scrollable_frame, text="Register", command=lambda: register_product(username, entry_names, labels, start_number, company_name, raw_labels, dropdown_vars))
    button_register_product.pack(padx=10, pady=30, anchor='center')

def register_product(username, entry_names, labels, start_number, company_name, raw_labels, dropdown_vars):
    df_company = pd.read_excel(resource_path('data//company_data.xlsx'))
    index = df_company.index[df_company['Username'] == username].tolist()
    row = df_company[df_company['Username'] == username]
    value = row.iat[0, row.columns.get_loc('Product')]

    product_name = entry_names[0].get().strip()
    product_dict = {}

    if len(product_name) != 0:
        count_label = 0
        for items in entry_names:
            product_dict[raw_labels[count_label]] = items.get().strip()
            count_label += 1

        for items in dropdown_vars:
            product_dict[raw_labels[count_label]] = items.get().strip()
            count_label += 1

        if pd.isna(value) or value == "":
            all_product_dict = {}
            all_product_dict[product_name] = product_dict
            df_company.iloc[index, 3] = str(all_product_dict)
            df_company.to_excel(resource_path('data//company_data.xlsx'), index=False)
        else:
            all_product_dict = eval(value)
            all_product_dict[product_name] = product_dict
            df_company.iloc[index, 3] = str(all_product_dict)
            df_company.to_excel(resource_path('data//company_data.xlsx'), index=False)

        
    else:
        messagebox.showerror("Product Registration", "Enter Product Name")

    prediction_dict = product_dict
    prediction_dict.pop('productName')

    discount = 15
    price = predict_asking_price(prediction_dict)
    workbook = openpyxl.load_workbook(resource_path('data//product_data.xlsx'))
    sheet = workbook.active
    new_data = [product_name, company_name, price, discount]
    sheet.append(new_data)
    workbook.save(resource_path('data//product_data.xlsx'))
    reload_company_dashboard(username, start_number)
# Function to handle registration
def register():
    new_username = entry_new_username.get().strip()
    new_password = entry_new_password.get().strip()
    confirm_password = entry_confirm_password.get().strip()
    new_name = entry_new_name.get().strip()
    user_type = selected_value.get().strip()

    if new_password == confirm_password:
        if user_type == "Customer":
            messagebox.showinfo("Registration", "Registration Successful!")
            workbook = openpyxl.load_workbook(resource_path('data//customer_data.xlsx'))
            sheet = workbook.active
            new_data = [new_username, new_password, new_name]
            sheet.append(new_data)
            workbook.save(resource_path('data//customer_data.xlsx'))
            entry_new_username.delete(0, tk.END)
            entry_new_password.delete(0, tk.END)
            entry_confirm_password.delete(0, tk.END)
            entry_new_name.delete(0, tk.END)
            show_login()
        else:
            messagebox.showinfo("Registration", "Registration Successful!")
            workbook = openpyxl.load_workbook(resource_path('data//company_data.xlsx'))
            sheet = workbook.active
            new_data = [new_username, new_password, new_name]
            sheet.append(new_data)
            workbook.save(resource_path('data//company_data.xlsx'))
            entry_new_username.delete(0, tk.END)
            entry_new_password.delete(0, tk.END)
            entry_confirm_password.delete(0, tk.END)
            show_login()
    else:
        messagebox.showerror("Registration", "Passwords do not match!")
        entry_confirm_password.delete(0, tk.END)

# Function to show login widgets
def show_login(event=None):
    # Hide registration widgets
    label_new_username.place_forget()
    entry_new_username.place_forget()
    label_new_password.place_forget()
    entry_new_password.place_forget()
    label_confirm_password.place_forget()
    entry_confirm_password.place_forget()
    button_register.place_forget()
    button_switch_to_login.place_forget()
    text_name_login.place_forget()
    text_label_login.place_forget()
    dropdown.place_forget()
    label_user_type.place_forget()
    label_new_name.place_forget()
    entry_new_name.place_forget()

    # Show login widgets
    label_username.place(x=250, y=150)
    entry_username.place(x=250, y=170, width=200, height=20)
    label_password.place(x=250, y=200)
    entry_password.place(x=250, y=220, width=200, height=20)
    button_login.place(x=300, y=260, width=100, height=25)
    text_name.place(x=250, y=300)
    text_label.place(x=310, y=299)
    dropdown_login.place(x=325, y=110)
    label_user_type_login.place(x=250, y=118)

# Function to show registration widgets
def show_register(event=None):
    # Hide login widgets
    label_username.place_forget()
    entry_username.place_forget()
    label_password.place_forget()
    entry_password.place_forget()
    button_login.place_forget()
    text_label.place_forget()
    text_name.place_forget()
    dropdown_login.place_forget()
    label_user_type_login.place_forget()
    # Show registration widgets
    label_new_username.place(x=250, y=180)
    entry_new_username.place(x=250, y=200, width=200, height=20)
    label_new_password.place(x=250, y=230)
    entry_new_password.place(x=250, y=250, width=200, height=20)
    label_confirm_password.place(x=250, y=280)
    entry_confirm_password.place(x=250, y=300, width=200, height=20)
    button_register.place(x=300, y=340, width=100, height=25)
    text_label_login.place(x=360, y=379)
    text_name_login.place(x=250, y=380)
    dropdown.place(x=325, y=85)
    label_user_type.place(x=250, y=93)
    label_new_name.place(x=250, y=130)
    entry_new_name.place(x=250, y=150, width=200, height=20)

def preprocess_input(user_input):
    """
    Convert user input into the same format as the training data.

    Args:
    user_input (dict): A dictionary containing user inputs.

    Returns:
    pd.DataFrame: A DataFrame ready to be input into the model for prediction.
    """




    original_columns = ['revenueMultiple', 'totalRevenueAnnual', 'team',
       'revenue', 'annualProfit', 'growthAnnual', 'weeklyViews', 'businessAge',
       'location_Argentina', 'location_Australia', 'location_Austria',
       'location_Belgium', 'location_Canada', 'location_Estonia',
       'location_France', 'location_Germany', 'location_India',
       'location_Latvia', 'location_Mexico', 'location_Singapore',
       'location_Spain', 'location_Switzerland', 'location_Turkey',
       'location_Ukraine', 'location_United Kingdom', 'location_United States',
       'keywords_Analytics', 'keywords_Automation', 'keywords_E-Commerce',
       'keywords_EdTech', 'keywords_Enterprise', 'keywords_Entertainment',
       'keywords_Finance', 'keywords_FundRaiser', 'keywords_Games',
       'keywords_HRTech', 'keywords_HealthCare', 'keywords_Loyalty',
       'keywords_Management', 'keywords_Marketing', 'keywords_Mobile App',
       'keywords_SEO', 'keywords_Sales', 'keywords_Security',
       'keywords_Servers', 'keywords_Technology', 'keywords_Telecommunication',
       'keywords_Web Tools', 'keywords_Workforce ',
       'techStack_ Cloud and DBMS', 'techStack_BackEnd',
       'techStack_BackEnd and Cloud',
       'techStack_Backend and DBMS', 'techStack_Cloud',
       'techStack_CyberSecurity', 'techStack_DBMS', 'techStack_FrontEnd',
       'techStack_FrontEnd and Cloud', 'techStack_FullStack',
       'techStack_FullStack and Cloud', 'techStack_FullStack and DBMS',
       'techStack_Not Listed', 'listingType_platinum', 'listingType_premium',
       'customers_10-100', 'customers_100-1000', 'customers_1000-10000',
       'customers_10000-100000', 'customers_more than 100000']

    # Initialize an empty DataFrame with the same columns as the training set
    input_data = pd.DataFrame(columns=original_columns)

    # Fill numeric inputs directly
    for col in ['revenueMultiple', 'totalRevenueAnnual', 'team', 'revenue', 'annualProfit', 'growthAnnual', 'weeklyViews','businessAge']:
        input_data.at[0, col] = user_input.get(col, 0)

    # Handle categorical inputs and convert to one-hot encoding
    # Reset all possible one-hot encoded categorical columns to 0
    for col in original_columns:
        if col.startswith('location_') or col.startswith('keywords_') or col.startswith('techStack_') or col.startswith('listingType_') or col.startswith('customers_'):
            input_data.at[0, col] = 0

    # Set the appropriate column to 1 for each categorical input
    if 'location' in user_input:
        input_data.at[0, f"location_{user_input['location']}"] = 1
    if 'keywords' in user_input:
        input_data.at[0, f"keywords_{user_input['keywords']}"] = 1
    if 'techStack' in user_input:
        input_data.at[0, f"techStack_{user_input['techStack']}"] = 1
    if 'listingType' in user_input:
        input_data.at[0, f"listingType_{user_input['listingType']}"] = 1
    if 'customers' in user_input:
        input_data.at[0, f"customers_{user_input['customers']}"] = 1

    # Convert all data to numeric types
    input_data = input_data.apply(pd.to_numeric, errors='coerce')

    return input_data

def predict_asking_price(user_input):
    """
    Predict the asking price using the trained XGBoost model.

    Args:
    user_input (dict): A dictionary containing input data for a single sample.

    Returns:
    float: The predicted asking price.
    """
    # Preprocess input data
    processed_input = preprocess_input(user_input)

    # Predict using the trained model
    predicted_price_log = xgb_model.predict(processed_input)
    # Inverse log transformation to get the original scale
    # predicted_price = np.expm1(predicted_price_log)

    return predicted_price_log[0]

# Create the main window
window = tk.Tk()
window.title("Login and Registration")
window.geometry("700x500")

# Login elements
label_username = tk.Label(window, text="Username:")
entry_username = tk.Entry(window)
label_password = tk.Label(window, text="Password:")
entry_password = tk.Entry(window, show="*")
button_login = tk.Button(window, text="Login", command=login)
text_name = tk.Label(window, text="New User?")

selected_value_login = tk.StringVar()
selected_value_login.set("Customer")
options_login = ["Customer", "Company"]
dropdown_login = tk.OptionMenu(window, selected_value_login, *options_login)

label_user_type_login = tk.Label(window, text="Log in as")

# Registration elements
label_new_username = tk.Label(window, text="New Username:")
label_new_name = tk.Label(window, text="Registered Name:")
entry_new_name = tk.Entry(window)
entry_new_username = tk.Entry(window)
label_new_password = tk.Label(window, text="New Password:")
entry_new_password = tk.Entry(window, show="*")
label_confirm_password = tk.Label(window, text="Confirm Password:")
entry_confirm_password = tk.Entry(window, show="*")
button_register = tk.Button(window, text="Register", command=register)
text_name_login = tk.Label(window, text="Already Registered?")

selected_value = tk.StringVar()
selected_value.set("Customer")
options = ["Customer", "Company"]
dropdown = tk.OptionMenu(window, selected_value, *options)

label_user_type = tk.Label(window, text="Sign up as")

button_switch_to_register = tk.Button(window, text="Switch to Register", command=show_register)
button_switch_to_login = tk.Button(window, text="Switch to Login", command=show_login)

text_label = tk.Label(window, text="Sign Up", fg="blue", cursor="hand2", font=("Helvetica", 9, "underline"))
text_label.bind("<Button-1>", show_register)

text_label_login = tk.Label(window, text="Log in", fg="blue", cursor="hand2", font=("Helvetica", 9, "underline"))
text_label_login.bind("<Button-1>", show_login)

show_login()

window.mainloop()
